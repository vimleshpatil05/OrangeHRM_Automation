import  openpyxl


class XLUtils_class:
    @staticmethod
    def get_row_count(file_path, sheet_name):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.max_row


    @staticmethod
    def read_data(file_path, sheet_name, row_num, col_num):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.cell(row=row_num, column=col_num).value

    @staticmethod
    def write_data(file_path, sheet_name, row_num, col_num, data):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        sheet.cell(row=row_num, column=col_num).value = data
        workbook.save(file_path)
        workbook.close()
